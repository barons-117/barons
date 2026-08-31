#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BARONS — המרת גיליון החיסכון והביטוחים ל-JSON עבור מחולל החוברת.

    python3 savings_from_xlsx.py Barons_Savings.xlsx -o savings.json

הגיליון מגיע ידנית מפרויקט 'פיננסי וביטוחים'. הממיר עמיד לשינויים קלים:
העמודות מזוהות לפי הכותרת בשורה 1, לא לפי מיקום. עמודה שנעלמה פשוט נשמטת.
"""

import argparse, json, sys
import openpyxl

# לכל גיליון: (כותרות להצגה כטבלה, כותרות שיורדות לשורת ההערות)
SHEETS = {
    "חסכונות ופנסיה": {
        "cols":  ["קטגוריה", "על שם", "חברה", "מס' פוליסה", "יתרה (₪)",
                  "הפקדה אחרונה (₪)", "דמי ניהול", "נזילות", "סטטוס"],
        "notes": ["הערות", "מוטבים", "סוכן"],
        "money": ["יתרה (₪)", "הפקדה אחרונה (₪)"],
    },
    "ביטוחי סיכון ובריאות": {
        "cols":  ["סוג כיסוי", "מבוטח", "חברה", "מס' פוליסה", "סכום ביטוח (₪)",
                  "פרמיה חודשית (₪)", "תום תקופה", "סטטוס"],
        "notes": ["הערות", "מוטבים", "סוכן"],
        "money": ["סכום ביטוח (₪)", "פרמיה חודשית (₪)"],
    },
    "רכוש - דירות ורכב": {
        "cols":  ["סוג", "נכס", "חברה", "מס' פוליסה", "סכום ביטוח (₪)",
                  "פרמיה שנתית (₪)", "תקופת ביטוח", "חידוש הבא", "סטטוס"],
        "notes": ["הערות", "סוכן"],
        "money": ["סכום ביטוח (₪)", "פרמיה שנתית (₪)"],
    },
    "חסכונות והשקעות אחרות": {
        "cols":  ["סוג ההשקעה\\חסכון", "מהות ההשקעה\\חסכון", "חברה2",
                  "שוי חסכון", "ערך", "תשואה חודשית", "תאריך עדכון"],
        "notes": ["הערות"],
        "money": [],
    },
    "סיכום אחזקות": {
        "cols":  ["קטגוריה", "פירוט", "סכום (₪)"],
        "notes": ["הערה"],
        "money": ["סכום (₪)"],
    },
    "שינויים בתהליך": {
        "cols":  ["תחום", "פוליסה/מוצר", "השינוי המתוכנן", "אחראי", "סטטוס", "נפתח"],
        "notes": ["בוצע? פרמיה/תוצאה"],
        "money": [],
    },
}

# גיליון היסטורי — כפילות של הנתונים החיים, לא נכנס לחוברת
SKIP = {"ישן לא פעיל"}

SKIP_ROW_PREFIXES = ("מקרא:",)


def cell(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def fmt_money(v):
    try:
        n = float(str(v).replace(",", "").replace("₪", "").strip())
    except (ValueError, TypeError):
        return cell(v)
    if n == 0:
        return "0"
    return f"{n:,.2f}".rstrip("0").rstrip(".") if n % 1 else f"{int(n):,}"


def convert(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sections = []

    for name, cfg in SHEETS.items():
        if name not in wb.sheetnames or name in SKIP:
            print(f"  דילוג — הגיליון '{name}' לא נמצא", file=sys.stderr)
            continue
        ws = wb[name]

        header = [cell(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        idx = {h: i for i, h in enumerate(header) if h}

        cols  = [c for c in cfg["cols"]  if c in idx]
        notes = [c for c in cfg["notes"] if c in idx]
        missing = [c for c in cfg["cols"] + cfg["notes"] if c not in idx]
        if missing:
            print(f"  ⚠ '{name}': עמודות חסרות — {', '.join(missing)}", file=sys.stderr)

        rows, total = [], None
        for raw in ws.iter_rows(min_row=2, values_only=True):
            vals = [cell(c) for c in raw]
            if not any(vals):
                continue
            first = vals[0] if vals else ""
            if first.startswith(SKIP_ROW_PREFIXES):
                continue
            # שורות המקרא: טקסט בעמודה הראשונה בלבד, כל השאר ריק.
            # ⚠️ אי אפשר לזהות אותן לפי הרווחים המובילים — cell() כבר עשה strip().
            if sum(1 for v in vals if v) == 1 and first not in ('סה"כ', "סה״כ"):
                continue

            get = lambda c: vals[idx[c]] if idx[c] < len(vals) else ""

            if first in ('סה"כ', "סה״כ", "סה\"כ שווי נקי", "סך הכל"):
                total = [(c, fmt_money(get(c)) if c in cfg["money"] else get(c))
                         for c in cols]
                continue

            cells = []
            for c in cols:
                v = get(c)
                cells.append(fmt_money(v) if c in cfg["money"] and v else v)
            if not any(cells):
                continue

            note_parts = []
            for c in notes:
                v = get(c)
                if v and v not in ("-", "?"):
                    note_parts.append(v if c == "הערות" else f"{c}: {v}")

            rows.append({"cells": cells, "notes": " · ".join(note_parts)})

        if rows:
            sections.append({"title": name, "columns": cols,
                             "rows": rows, "total": total})
            print(f"  {name}: {len(rows)} שורות", file=sys.stderr)

    return {"sections": sections}


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("-o", "--out", default="savings.json")
    a = ap.parse_args()
    data = convert(a.xlsx)
    json.dump(data, open(a.out, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"נוצר: {a.out}  ({len(data['sections'])} גיליונות)")
